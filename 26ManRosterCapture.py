"""
MLB Roster Fetcher
==================
Pulls the current 26-man active roster + IL players from the 40-man roster
for every MLB team, then writes one Excel tab per team.

Requirements:
    pip install requests openpyxl

Usage:
    python mlb_roster_fetcher.py
    python mlb_roster_fetcher.py --season 2025          # specific season
    python mlb_roster_fetcher.py --out my_rosters.xlsx  # custom filename
"""

import argparse
import datetime
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_URL = "https://statsapi.mlb.com/api/v1"
HEADERS = {"User-Agent": "mlb-roster-fetcher/1.0"}
CURRENT_SEASON = datetime.date.today().year

# Roster type codes used by the MLB Stats API
ACTIVE_ROSTER_TYPE = "active"      # 26-man active roster
FULL_ROSTER_TYPE   = "fullRoster"  # full 40-man (includes IL players)

# Column definitions: (header label, player-dict key, width)
COLUMNS = [
    ("Name",          "fullName",        28),
    ("Jersey #",      "jerseyNumber",     9),
    ("Position",      "position",        10),
    ("IL Type",       "ilType",          12),
    ("Inn-C",         "inn_C",            8),
    ("Inn-1B",        "inn_1B",           8),
    ("Inn-2B",        "inn_2B",           8),
    ("Inn-SS",        "inn_SS",           8),
    ("Inn-3B",        "inn_3B",           8),
    ("Inn-LF",        "inn_LF",           8),
    ("Inn-CF",        "inn_CF",           8),
    ("Inn-RF",        "inn_RF",           8),
    ("Inn-OF",        "inn_OF",           8),
    ("Inn-DH",        "inn_DH",           8),
    ("Inn-P",         "inn_P",            8),
    ("Bats",          "batSide",          7),
    ("Throws",        "pitchHand",        8),
    ("DOB",           "birthDate",       12),
    ("Birth Country", "birthCountry",    16),
    ("MLB Debut",     "mlbDebutDate",    12),
    ("Height",        "height",           8),
    ("Weight",        "weight",           8),
    ("Player ID",     "id",              10),
]

# Colour palette
CLR_HEADER_ACTIVE = "1A3A5C"   # dark navy  – 26-man header
CLR_HEADER_IL     = "8B1A1A"   # dark red   – IL header
CLR_ROW_ACTIVE    = "DDEEFF"   # light blue – 26-man rows (alternating)
CLR_ROW_IL        = "FFE8E8"   # light red  – IL rows (alternating)
CLR_WHITE         = "FFFFFF"

# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def _get(url: str, params=None) -> dict:
    resp = requests.get(url, params=params, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    return resp.json()


def fetch_teams(season: int) -> list[dict]:
    """Return all MLB teams sorted by name."""
    data = _get(f"{BASE_URL}/teams", params={"sportId": 1, "season": season})
    teams = data.get("teams", [])
    return sorted(teams, key=lambda t: t["name"])


def fetch_active_roster(team_id: int, season: int) -> list[dict]:
    """Return the 26-man active roster for a team."""
    try:
        data = _get(
            f"{BASE_URL}/teams/{team_id}/roster",
            params={"rosterType": ACTIVE_ROSTER_TYPE, "season": season},
        )
        return data.get("roster", [])
    except requests.HTTPError:
        return []


def fetch_full_roster(team_id: int, season: int) -> list[dict]:
    try:
        data = _get(
            f"{BASE_URL}/teams/{team_id}/roster",
            params={"rosterType": "40Man", "season": season},
        )
        return data.get("roster", [])
    except requests.HTTPError:
        return []


def fetch_player_detail(player_id: int) -> dict:
    """Return biographical data for a player."""
    try:
        data = _get(f"{BASE_URL}/people/{player_id}")
        people = data.get("people", [])
        return people[0] if people else {}
    except requests.HTTPError:
        return {}

def fetch_position_breakdown(player_id: int, season: int) -> dict:
    """
    Returns dict with:
      'primary'   : str  — position with most innings (DH+OF grouped for OF players)
      'breakdown' : dict — { 'SS': 91.0, '3B': 90.0, 'DH': 12 (games), ... }
    """
    breakdown = {}

    # ── Fielding innings by position ──────────────────────────────────────
    try:
        data = _get(
            f"{BASE_URL}/people/{player_id}/stats",
            params={"stats": "season", "group": "fielding", "season": season}
        )
        for stat_group in data.get("stats", []):
            for split in stat_group.get("splits", []):
                pos = split.get("position", {}).get("abbreviation", "")
                inn = split.get("stat", {}).get("innings", "0")
                if not pos:
                    continue
                # Convert MLB innings format: 91.1 = 91 and 1/3, 91.2 = 91 and 2/3
                try:
                    raw = str(inn)
                    parts = raw.split(".")
                    whole = float(parts[0])
                    frac  = int(parts[1]) / 3 if len(parts) > 1 else 0
                    breakdown[pos] = breakdown.get(pos, 0) + whole + frac
                except (ValueError, IndexError):
                    pass
    except requests.HTTPError:
        pass

   # ── DH innings estimate from fielding stats (games x 9) ──────────────
    try:
        data = _get(
            f"{BASE_URL}/people/{player_id}/stats",
            params={"stats": "season", "group": "fielding", "season": season}
        )
        for stat_group in data.get("stats", []):
            for split in stat_group.get("splits", []):
                pos = split.get("position", {}).get("abbreviation", "")
                if pos == "DH":
                    games = split.get("stat", {}).get("gamesPlayed", 0)
                    if games:
                        breakdown["DH"] = round(int(games) * 9, 1)
    except requests.HTTPError:
        pass

    # ── Determine primary position ────────────────────────────────────────
    if not breakdown:
        return {"primary": "", "breakdown": {}}

    fielding_only = {k: v for k, v in breakdown.items() if k != "DH"}
    
    # DH innings equivalent for comparison (games x 9)
    dh_inn_equiv = breakdown.get("DH", 0)  # already stored as games x 9

    of_positions = {"LF", "CF", "RF", "OF"}

    if not fielding_only:
        # Pure DH — group into OF
        primary = "OF"
    else:
        top_fielding     = max(fielding_only, key=fielding_only.get)
        top_fielding_inn = fielding_only.get(top_fielding, 0)

        if top_fielding in of_positions:
            # OF player — stays OF regardless of DH
            primary = "OF"
        elif dh_inn_equiv > top_fielding_inn:
            # More time as DH than any fielding position — group into OF
            primary = "OF"
        else:
            primary = top_fielding

    return {"primary": primary, "breakdown": breakdown}

# ---------------------------------------------------------------------------
# Roster processing
# ---------------------------------------------------------------------------

def player_row(roster_entry: dict, detail: dict, status_label: str, pos_data: dict = None) -> dict:
    pos_data    = pos_data or {}
    primary     = pos_data.get("primary", "")
    breakdown   = pos_data.get("breakdown", {})
    code        = roster_entry.get("status", {}).get("code", "")
    api_pos     = roster_entry.get("position", {}).get("abbreviation", "")
    
    il_labels = {
        "D7":  "7-Day IL", "D10": "10-Day IL",
        "D15": "15-Day IL", "D60": "60-Day IL", "ILF": "Full Season IL"
    }

    def fmt_inn(pos):
        val = breakdown.get(pos, "")
        if val == "":
            return ""
        return round(float(val), 1)

    return {
        "fullName":     detail.get("fullName", roster_entry.get("person", {}).get("fullName", "")),
        "jerseyNumber": roster_entry.get("jerseyNumber", ""),
        "position":     primary if primary else api_pos,
        "ilType":       il_labels.get(code, "") if status_label == "IL" else "",
        "inn_C":        fmt_inn("C"),
        "inn_1B":       fmt_inn("1B"),
        "inn_2B":       fmt_inn("2B"),
        "inn_SS":       fmt_inn("SS"),
        "inn_3B":       fmt_inn("3B"),
        "inn_LF":       fmt_inn("LF"),
        "inn_CF":       fmt_inn("CF"),
        "inn_RF":       fmt_inn("RF"),
        "inn_OF":       fmt_inn("OF"),
        "inn_DH":       fmt_inn("DH"),
        "inn_P":        fmt_inn("P"),
        "batSide":      detail.get("batSide", {}).get("code", ""),
        "pitchHand":    detail.get("pitchHand", {}).get("code", ""),
        "birthDate":    detail.get("birthDate", ""),
        "birthCountry": detail.get("birthCountry", ""),
        "mlbDebutDate": detail.get("mlbDebutDate", ""),
        "height":       detail.get("height", ""),
        "weight":       detail.get("weight", ""),
        "id":           detail.get("id", roster_entry.get("person", {}).get("id", "")),
    }


def get_team_roster(team_id: int, season: int):
    active_entries = fetch_active_roster(team_id, season)
    full_entries   = fetch_full_roster(team_id, season)

    active_ids = {e["person"]["id"] for e in active_entries}

    IL_CODES = {"D7", "D10", "D15", "D60", "ILF"}

    il_entries = [
        e for e in full_entries
        if e["person"]["id"] not in active_ids
        and e.get("status", {}).get("code", "") in IL_CODES
    ]

    def enrich(entries, status_label):
        rows = []
        for entry in entries:
            pid      = entry["person"]["id"]
            detail   = fetch_player_detail(pid)
            pos_data = fetch_position_breakdown(pid, season)
            time.sleep(0.05)
            rows.append(player_row(entry, detail, status_label, pos_data))
        return rows

    active_rows = enrich(active_entries, "Active")
    il_rows     = enrich(il_entries,     "IL")
    return active_rows, il_rows

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _hex(hex_str: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_str, fgColor=hex_str)


def _write_section_header(ws, row: int, label: str, color: str, num_cols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = _hex(color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18


def _write_column_headers(ws, row: int, color: str):
    for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _hex(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 16


def _write_player_rows(ws, start_row: int, players: list[dict], row_color: str):
    for i, player in enumerate(players):
        r = start_row + i
        bg = row_color if i % 2 == 0 else CLR_WHITE
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=r, column=col_idx, value=player.get(key, ""))
            cell.fill      = _hex(bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font      = Font(size=10)
    return start_row + len(players)


def build_team_sheet(wb: Workbook, team_name: str, active: list[dict], il: list[dict]):
    # Sanitise sheet name (max 31 chars, no special chars)
    safe_name = team_name.replace("/", "-").replace("\\", "-").replace("*", "")[:31]
    ws = wb.create_sheet(title=safe_name)
    ws.freeze_panes = "A3"

    num_cols = len(COLUMNS)
    cur_row  = 1

    # ── 26-Man Active Roster ──────────────────────────────────────────────
    _write_section_header(ws, cur_row, f"26-Man Active Roster  ({len(active)} players)",
                          CLR_HEADER_ACTIVE, num_cols)
    cur_row += 1
    _write_column_headers(ws, cur_row, CLR_HEADER_ACTIVE)
    cur_row += 1

    if active:
        POS_ORDER = ["P", "C", "1B", "2B", "SS", "3B", "OF"]
        active_sorted = sorted(
            active,
            key=lambda p: (
                POS_ORDER.index(p.get("position", "OF"))
                if p.get("position", "") in POS_ORDER
                else len(POS_ORDER),
                p.get("fullName", "")
            )
        )
        cur_row = _write_player_rows(ws, cur_row, active_sorted, CLR_ROW_ACTIVE)
    else:
        ws.cell(row=cur_row, column=1, value="No active roster data available")
        cur_row += 1

    # ── IL Players ────────────────────────────────────────────────────────
    _write_section_header(ws, cur_row, f"Injured List (40-man)  ({len(il)} players)",
                          CLR_HEADER_IL, num_cols)
    cur_row += 1
    _write_column_headers(ws, cur_row, CLR_HEADER_IL)
    cur_row += 1

    if il:
        POS_ORDER = ["P", "C", "1B", "2B", "SS", "3B", "OF"]
        il_sorted = sorted(
            il,
            key=lambda p: (
                POS_ORDER.index(p.get("position", "OF"))
                if p.get("position", "") in POS_ORDER
                else len(POS_ORDER),
                p.get("fullName", "")
            )
        )
        _write_player_rows(ws, cur_row, il_sorted, CLR_ROW_IL)
    else:
        ws.cell(row=cur_row, column=1, value="No IL players")

    # Auto-filter on column headers (row 2 for active section)
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"

    return ws


def build_summary_sheet(wb: Workbook, summary_rows: list[dict]):
    ws = wb.create_sheet(title="Summary", index=0)
    headers = ["Team", "Active (26-man)", "IL Players", "Total on Roster"]
    col_widths = [30, 16, 14, 16]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _hex(CLR_HEADER_ACTIVE)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for i, row in enumerate(summary_rows, 2):
        bg = CLR_ROW_ACTIVE if i % 2 == 0 else CLR_WHITE
        ws.cell(row=i, column=1, value=row["team"]).fill = _hex(bg)
        ws.cell(row=i, column=2, value=row["active"]).fill = _hex(bg)
        ws.cell(row=i, column=3, value=row["il"]).fill = _hex(bg)
        ws.cell(row=i, column=4, value=row["active"] + row["il"]).fill = _hex(bg)

    # Totals row
    last = len(summary_rows) + 2
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last, column=2, value=f"=SUM(B2:B{last-1})").font = Font(bold=True)
    ws.cell(row=last, column=3, value=f"=SUM(C2:C{last-1})").font = Font(bold=True)
    ws.cell(row=last, column=4, value=f"=SUM(D2:D{last-1})").font = Font(bold=True)

    ws.auto_filter.ref = f"A1:D{last-1}"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(description="Fetch MLB rosters to Excel")
    p.add_argument("--season", type=int, default=CURRENT_SEASON,
                   help=f"MLB season year (default: {CURRENT_SEASON})")
    p.add_argument("--out", default=r"C:\FantasyBaseballSite\26ManRosters.xlsx",
                   help="Output Excel filename (default: mlb_rosters.xlsx)")
    p.add_argument("--team", type=str, default=None,
                   help="Filter to a single team name (partial match, case-insensitive)")
    return p.parse_args()


def main():
    args = parse_args()
    season = args.season

    print(f"Fetching MLB teams for {season}...")
    teams = fetch_teams(season)

    if args.team:
        filter_str = args.team.lower()
        teams = [t for t in teams if filter_str in t["name"].lower()]
        if not teams:
            print(f"No teams matched '{args.team}'")
            sys.exit(1)
        print(f"Filtered to {len(teams)} team(s): {[t['name'] for t in teams]}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    summary_rows = []
    total = len(teams)

    for idx, team in enumerate(teams, 1):
        team_id   = team["id"]
        team_name = team["name"]
        print(f"[{idx:2}/{total}] {team_name}...", end=" ", flush=True)

        active, il = get_team_roster(team_id, season)
        print(f"{len(active)} active, {len(il)} IL")

        build_team_sheet(wb, team_name, active, il)
        summary_rows.append({"team": team_name, "active": len(active), "il": len(il)})

    build_summary_sheet(wb, summary_rows)

    wb.save(args.out)
    print(f"\n✅  Saved → {args.out}")
    print(f"   {total} team tabs + 1 Summary tab")
    total_active = sum(r["active"] for r in summary_rows)
    total_il     = sum(r["il"]     for r in summary_rows)
    print(f"   {total_active} active players  |  {total_il} IL players")


if __name__ == "__main__":
    main()